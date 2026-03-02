import frappe
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import Font
from frappe.utils.xlsxutils import make_xlsx

@frappe.whitelist()
def export():
    result = frappe.call(
        "general_ledger_customizations.api.balance_sheet_data.get_balance_sheet_prepared_data"
    )

    if not result.get("success"):
        frappe.throw(result.get("message"))

    columns = result["columns"]   # list of dicts
    data = result["data"]         # list of dicts

    rows = []

    # Header row (labels)
    rows.append([col.get("label") for col in columns])

    # Extract values in column order
    fieldnames = [col.get("fieldname") for col in columns]

    for row in data:
        rows.append([row.get(field) for field in fieldnames])

    # Create Excel
    xlsx_file = make_xlsx(rows, "Balance Sheet")

    wb = load_workbook(BytesIO(xlsx_file.getvalue()))
    ws = wb.active

    from openpyxl.styles import Font

    bold = Font(bold=True)
    normal = Font(bold=False)

    # Header always bold (entire header row)
    for col in range(1, ws.max_column + 1):
        ws.cell(row=1, column=col).font = bold

    # Conditional bolding for entire rows
    for excel_row_index, row_data in enumerate(data, start=2):
        account_name = str(row_data.get("account") or "").strip()

        is_parent = row_data.get("is_group") == 1
        is_total = "total" in account_name.lower()
        is_profit_loss = "profit" in account_name.lower() or "loss" in account_name.lower()

        # Decide font for the row
        row_font = bold if (is_parent or is_total or is_profit_loss) else normal

        # Apply font to the ENTIRE row
        for col in range(1, ws.max_column + 1):
            ws.cell(row=excel_row_index, column=col).font = row_font

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    frappe.response["filename"] = "Balance Sheet.xlsx"
    frappe.response["filecontent"] = output.read()
    frappe.response["type"] = "binary"